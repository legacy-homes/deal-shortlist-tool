#!/usr/bin/env python3.11
"""
EPC Floor Area Extraction Tool
Extracts total floor area from UK Energy Performance Certificate database
"""

import requests
import re
import time
import argparse
import json
from bs4 import BeautifulSoup
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font

def normalize_address(address):
    """Normalize address for comparison"""
    # Convert to uppercase, remove extra spaces, standardize punctuation
    normalized = address.upper().strip()
    normalized = re.sub(r'\s+', ' ', normalized)
    normalized = re.sub(r',\s*', ', ', normalized)
    return normalized

def similarity_score(addr1, addr2):
    """Calculate similarity score between two addresses"""
    norm1 = normalize_address(addr1)
    norm2 = normalize_address(addr2)
    return SequenceMatcher(None, norm1, norm2).ratio()

def extract_postcode_from_address(address):
    """Extract postcode from full address"""
    # UK postcode pattern
    postcode_match = re.search(r'\b([A-Z]{1,2}\d{1,2}[A-Z]?)\s*(\d[A-Z]{2})\b', address.upper())
    if postcode_match:
        return postcode_match.group(1) + ' ' + postcode_match.group(2)
    return None

def search_epc_by_postcode(postcode):
    """Search for EPCs by postcode"""
    url = "https://find-energy-certificate.service.gov.uk/find-a-certificate/search-by-postcode"
    params = {
        'postcode': postcode
    }
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    try:
        response = requests.get(url, params=params, headers=headers, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Find all address links in the results
        addresses = []
        for link in soup.find_all('a', href=re.compile(r'/energy-certificate/')):
            address_text = link.get_text(strip=True)
            certificate_url = link['href']
            
            # Extract certificate number from URL
            cert_match = re.search(r'/energy-certificate/(.+)$', certificate_url)
            if cert_match:
                cert_number = cert_match.group(1)
                addresses.append({
                    'address': address_text,
                    'certificate_number': cert_number,
                    'url': 'https://find-energy-certificate.service.gov.uk' + certificate_url
                })
        
        return addresses
    
    except requests.RequestException as e:
        print(f"    Error searching EPC: {e}")
        return []

def get_floor_area_from_certificate(certificate_url):
    """Get floor area from EPC certificate page"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    try:
        response = requests.get(certificate_url, headers=headers, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Look for "Total floor area" text
        floor_area_text = soup.find(string=re.compile(r'Total floor area', re.IGNORECASE))
        
        if floor_area_text:
            # Find the parent element and look for the value
            parent = floor_area_text.find_parent()
            if parent:
                # Look for the next sibling or nearby text with square metres
                for sibling in parent.find_next_siblings():
                    text = sibling.get_text(strip=True)
                    # Match patterns like "119 square metres" or "119 sq m"
                    match = re.search(r'(\d+(?:\.\d+)?)\s*(?:square\s*metres?|sq\s*m)', text, re.IGNORECASE)
                    if match:
                        return float(match.group(1))
                
                # Also check in the same element
                text = parent.get_text(strip=True)
                match = re.search(r'Total floor area[:\s]*(\d+(?:\.\d+)?)\s*(?:square\s*metres?|sq\s*m)', text, re.IGNORECASE)
                if match:
                    return float(match.group(1))
        
        # Alternative: Look in table/definition list format
        for dt in soup.find_all(['dt', 'th']):
            if 'floor area' in dt.get_text(strip=True).lower():
                # Find the corresponding dd or td
                dd = dt.find_next_sibling(['dd', 'td'])
                if dd:
                    text = dd.get_text(strip=True)
                    match = re.search(r'(\d+(?:\.\d+)?)', text)
                    if match:
                        return float(match.group(1))
        
        return None
    
    except requests.RequestException as e:
        print(f"    Error fetching certificate: {e}")
        return None

def extract_house_number(address):
    """Extract house number from address"""
    # Match patterns like "29," or "29 " at the start
    match = re.match(r'^(\d+[A-Z]?)\s*[,\s]', address.upper().strip())
    if match:
        return match.group(1)
    return None

def find_floor_area_for_address(full_address, threshold=0.85):
    """
    Find floor area for a given address
    
    Args:
        full_address: Full address string
        threshold: Similarity threshold for address matching (0-1)
    
    Returns:
        dict with 'floor_area' and 'availability' keys
    """
    # Extract postcode from address
    postcode = extract_postcode_from_address(full_address)
    
    if not postcode:
        return {
            'floor_area': -1,
            'availability': 'NotAvailable',
            'reason': 'No postcode found in address'
        }
    
    # Search for EPCs with this postcode
    epc_addresses = search_epc_by_postcode(postcode)
    
    if not epc_addresses:
        return {
            'floor_area': -1,
            'availability': 'NotAvailable',
            'reason': 'No EPCs found for postcode'
        }
    
    # Extract house number from target address
    target_house_num = extract_house_number(full_address)
    
    # Find best matching address
    best_match = None
    best_score = 0
    
    for epc_addr in epc_addresses:
        # First check if house numbers match (if we have one)
        if target_house_num:
            epc_house_num = extract_house_number(epc_addr['address'])
            if epc_house_num and epc_house_num != target_house_num:
                # Skip if house numbers don't match
                continue
        
        score = similarity_score(full_address, epc_addr['address'])
        if score > best_score:
            best_score = score
            best_match = epc_addr
    
    # Check if match is good enough
    if best_score < threshold:
        return {
            'floor_area': -1,
            'availability': 'NotAvailable',
            'reason': f'No matching address found (best match: {best_score:.2f})',
            'best_match_address': best_match['address'] if best_match else None,
            'best_match_score': best_score
        }
    
    # Get floor area from certificate
    floor_area = get_floor_area_from_certificate(best_match['url'])
    
    if floor_area is not None:
        return {
            'floor_area': floor_area,
            'availability': 'Available',
            'matched_address': best_match['address'],
            'match_score': best_score,
            'certificate_url': best_match['url']
        }
    else:
        return {
            'floor_area': -1,
            'availability': 'NotAvailable',
            'reason': 'Floor area not found in certificate',
            'matched_address': best_match['address'],
            'certificate_url': best_match['url']
        }

def process_excel_file(input_file, output_file, address_column='Full Address', 
                       floor_area_column='Total Floor Area (sq m)', 
                       availability_column='Floor Area Availability',
                       sheet_name='SoldProperties'):
    """
    Process Excel file and add floor area data
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file
        address_column: Name of column containing addresses
        floor_area_column: Name of column to add floor area
        availability_column: Name of column to add availability status
        sheet_name: Name of sheet to process
    """
    print(f"Loading Excel file: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found")
        return
    
    ws = wb[sheet_name]
    
    # Find column indices
    headers = [cell.value for cell in ws[1]]
    
    if address_column not in headers:
        print(f"Error: Column '{address_column}' not found")
        return
    
    address_col_idx = headers.index(address_column) + 1
    
    # Add new columns if they don't exist
    if floor_area_column not in headers:
        floor_area_col_idx = len(headers) + 1
        ws.cell(1, floor_area_col_idx, floor_area_column).font = Font(bold=True)
    else:
        floor_area_col_idx = headers.index(floor_area_column) + 1
    
    if availability_column not in headers:
        availability_col_idx = len(headers) + 2 if floor_area_column not in headers else len(headers) + 1
        ws.cell(1, availability_col_idx, availability_column).font = Font(bold=True)
    else:
        availability_col_idx = headers.index(availability_column) + 1
    
    # Process each row
    total_rows = ws.max_row - 1
    print(f"\nProcessing {total_rows} properties...")
    print("=" * 80)
    
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row, address_col_idx).value
        
        if not address:
            continue
        
        print(f"[{row-1}/{total_rows}] {address[:60]}...", end=' ', flush=True)
        
        # Get floor area
        result = find_floor_area_for_address(address)
        
        # Update cells
        ws.cell(row, floor_area_col_idx, result['floor_area'])
        ws.cell(row, availability_col_idx, result['availability'])
        
        if result['availability'] == 'Available':
            print(f"✓ {result['floor_area']} sq m")
        else:
            print(f"✗ {result.get('reason', 'Not available')}")
        
        # Be polite to the server
        time.sleep(2)
    
    print("=" * 80)
    print(f"\nSaving results to: {output_file}")
    wb.save(output_file)
    print("✓ Complete!")

def main():
    parser = argparse.ArgumentParser(
        description='Extract floor area from UK EPC database for addresses in Excel file'
    )
    parser.add_argument('--input', help='Input Excel file')
    parser.add_argument('--output', help='Output Excel file')
    parser.add_argument('--sheet', default='SoldProperties', help='Sheet name (default: SoldProperties)')
    parser.add_argument('--address-column', default='Full Address', help='Address column name')
    parser.add_argument('--floor-area-column', default='Total Floor Area (sq m)', 
                       help='Floor area column name')
    parser.add_argument('--availability-column', default='Floor Area Availability',
                       help='Availability column name')
    parser.add_argument('--test-address', help='Test with a single address')
    
    args = parser.parse_args()
    
    if args.test_address:
        print(f"Testing with address: {args.test_address}")
        result = find_floor_area_for_address(args.test_address)
        print(json.dumps(result, indent=2))
    elif args.input and args.output:
        process_excel_file(
            args.input,
            args.output,
            address_column=args.address_column,
            floor_area_column=args.floor_area_column,
            availability_column=args.availability_column,
            sheet_name=args.sheet
        )
    else:
        parser.error('Either --test-address or both --input and --output are required')

if __name__ == '__main__':
    main()
