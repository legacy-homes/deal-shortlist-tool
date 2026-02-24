#!/usr/bin/env python3.11
"""
Property Deal Finder - Modular Version
=======================================

Extracts active property listings from Rightmove using the shared parsing module,
compares asking prices with median sold prices, and creates a shortlist of
potentially undervalued properties.

Version: 3.0.0 (Progressive Search)
Author: Manus AI
Date: 2026-02-14

Changes from v2.0:
- Uses MedianPriceCalculator v3.0.0 with progressive search
- Automatically expands search radius and time period to find enough properties
- Added tenure filter (FREEHOLD)
- More accurate median price calculations
"""

import requests
import sys
import os
import argparse
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Import shared parsing module
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'shared'))
from rightmove_parsers import extract_active_properties_from_html

# Import the median price calculator
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'MedianPriceCalculator'))
from median_price_calculator import calculate_median_price_progressive

class PropertyDealFinder:
    def __init__(self, config_file='config.json'):
        """Initialize the PropertyDealFinder with configuration."""
        self.config = self.load_config(config_file)
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
    def load_config(self, config_file):
        """Load configuration from JSON file."""
        if os.path.exists(config_file):
            import json
            with open(config_file, 'r') as f:
                return json.load(f)
        return {
            'price_difference_threshold': 50000,
            'radius_for_median': 0.25,
            'delay_between_requests': 2,
            'max_retries': 3
        }
    
    def fetch_page(self, url):
        """Fetch a page HTML."""
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return response.text
        except Exception as e:
            print(f"Error fetching page: {e}")
            return None
    
    def extract_properties(self, html):
        """
        Extract property listings from HTML using shared parser.
        
        NOTE: This now uses the shared rightmove_parsers module.
        When Rightmove changes their structure, only rightmove_parsers.py needs updating.
        """
        if not html:
            return []
        
        # Use shared parser - it returns standardized property dicts
        properties = extract_active_properties_from_html(html, include_featured=False)
        
        # Filter out None values (properties that failed validation)
        valid_properties = [p for p in properties if p is not None]
        
        return valid_properties
    
    def normalize_property_type(self, property_type):
        """Normalize property type for median calculation."""
        # Rightmove uses different property type names in listings vs sold data
        # Normalize to match the sold data format
        type_mappings = {
            'End of Terrace': 'Terraced',
            'End Terrace': 'Terraced',
            'Mid Terrace': 'Terraced',
            'Mid Terraced': 'Terraced',
            'Terraced House': 'Terraced',
            'Semi-Detached House': 'Semi-Detached',
            'Detached House': 'Detached',
        }
        return type_mappings.get(property_type, property_type)
    
    def compare_with_median(self, property_data):
        """Compare asking price with median sold price."""
        try:
            postcode = property_data['postcode']
            property_type = property_data['property_type']
            bedrooms = property_data['bedrooms']
            radius = self.config['radius_for_median']
            
            # Normalize property type for median calculation
            normalized_type = self.normalize_property_type(property_type)
            
            print(f"  Calculating median for {postcode}, {property_type}, {bedrooms} beds...")
            if normalized_type != property_type:
                print(f"  (Normalized to '{normalized_type}' for median calculation)")
            
            median_result = calculate_median_price_progressive(
                postcode=postcode,
                property_type=normalized_type,
                bedrooms=bedrooms,
                tenure='FREEHOLD',
                min_properties=5  # Use lower threshold for deal finder
            )
            
            if median_result and median_result['median_price']:
                median_price = median_result['median_price']
                sample_size = median_result['property_count']
                asking_price = property_data['price']
                difference = median_price - asking_price
                
                return {
                    'median_price': median_price,
                    'sample_size': sample_size,
                    'difference': difference,
                    'is_deal': difference >= self.config['price_difference_threshold']
                }
            else:
                print(f"  No median data available")
                return None
                
        except Exception as e:
            print(f"  Error calculating median: {e}")
            return None
    
    def create_shortlist_spreadsheet(self, deals, output_file):
        """Create Excel spreadsheet with shortlisted deals."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Deals"
        
        headers = [
            'Property ID', 'Postcode', 'Address', 'Property Type', 'Bedrooms',
            'Asking Price (£)', 'Median Price (£)', 'Difference (£)',
            'Sample Size', 'Listing Link'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, deal in enumerate(deals, 2):
            ws.cell(row=row_idx, column=1, value=deal['id'])
            ws.cell(row=row_idx, column=2, value=deal['postcode'])
            ws.cell(row=row_idx, column=3, value=deal['address'])
            ws.cell(row=row_idx, column=4, value=deal['property_type'])
            ws.cell(row=row_idx, column=5, value=deal['bedrooms'])
            ws.cell(row=row_idx, column=6, value=deal['price'])
            ws.cell(row=row_idx, column=7, value=deal['median_price'])
            ws.cell(row=row_idx, column=8, value=deal['difference'])
            ws.cell(row=row_idx, column=9, value=deal['sample_size'])
            ws.cell(row=row_idx, column=10, value=deal['link'])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        print(f"\nShortlist saved to: {output_file}")
    
    def process_listings(self, url, first_page_only=True, start_index=0, max_properties=None):
        """Process property listings and create shortlist."""
        deals = []
        page_index = start_index
        properties_processed = 0
        
        while True:
            page_url = url.replace('index=0', f'index={page_index}')
            
            print(f"\n{'='*80}")
            print(f"Fetching page {page_index//24 + 1} (index={page_index})...")
            print(f"{'='*80}")
            
            # Fetch HTML
            html = self.fetch_page(page_url)
            if not html:
                print("Failed to fetch page")
                break
            
            # Extract listings using shared parser
            listings = self.extract_properties(html)
            print(f"Found {len(listings)} valid properties on this page (excluding featured)")
            
            if not listings:
                print("No more properties found")
                break
            
            # Process each listing
            for listing in listings:
                properties_processed += 1
                
                if max_properties and properties_processed > max_properties:
                    print(f"\nReached maximum properties limit ({max_properties})")
                    break
                
                print(f"\n[{properties_processed}] {listing['address']}")
                print(f"  Asking Price: £{listing['price']:,}")
                
                # Compare with median
                comparison = self.compare_with_median(listing)
                
                if comparison:
                    print(f"  Median Price: £{comparison['median_price']:,} (based on {comparison['sample_size']} properties)")
                    print(f"  Difference: £{comparison['difference']:,}")
                    
                    if comparison['is_deal']:
                        print(f"  ✓ POTENTIAL DEAL! (Difference >= £{self.config['price_difference_threshold']:,})")
                        deals.append({
                            'id': listing['id'],
                            'postcode': listing['postcode'],
                            'address': listing['address'],
                            'property_type': listing['property_type'],
                            'bedrooms': listing['bedrooms'],
                            'price': listing['price'],
                            'median_price': comparison['median_price'],
                            'difference': comparison['difference'],
                            'sample_size': comparison['sample_size'],
                            'link': listing['link']
                        })
                    else:
                        print(f"  Not a deal (Difference < £{self.config['price_difference_threshold']:,})")
                else:
                    print(f"  Skipped (no median data)")
                
                # Delay between requests
                time.sleep(self.config['delay_between_requests'])
            
            if max_properties and properties_processed >= max_properties:
                break
            
            if first_page_only:
                break
            
            # Move to next page
            page_index += 24
            time.sleep(self.config['delay_between_requests'])
        
        return deals


def main():
    parser = argparse.ArgumentParser(
        description='Find potentially undervalued properties on Rightmove'
    )
    parser.add_argument(
        'url',
        help='Rightmove search URL'
    )
    parser.add_argument(
        '--config',
        default='config.json',
        help='Configuration file (default: config.json)'
    )
    parser.add_argument(
        '--output',
        default=None,
        help='Output Excel file (default: auto-generated)'
    )
    parser.add_argument(
        '--all-pages',
        action='store_true',
        help='Process all pages (default: first page only)'
    )
    parser.add_argument(
        '--max-properties',
        type=int,
        default=None,
        help='Maximum number of properties to process'
    )
    
    args = parser.parse_args()
    
    # Initialize finder
    finder = PropertyDealFinder(config_file=args.config)
    
    # Process listings
    print("="*80)
    print("PROPERTY DEAL FINDER - Modular Version 2.0.0")
    print("="*80)
    print(f"Search URL: {args.url}")
    print(f"Price Difference Threshold: £{finder.config['price_difference_threshold']:,}")
    print(f"Radius for Median: {finder.config['radius_for_median']} miles")
    print(f"Mode: {'All pages' if args.all_pages else 'First page only'}")
    if args.max_properties:
        print(f"Max Properties: {args.max_properties}")
    print("="*80)
    
    deals = finder.process_listings(
        url=args.url,
        first_page_only=not args.all_pages,
        max_properties=args.max_properties
    )
    
    # Create output
    if deals:
        if not args.output:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            args.output = f'property_deals_{timestamp}.xlsx'
        
        finder.create_shortlist_spreadsheet(deals, args.output)
        
        print(f"\n{'='*80}")
        print(f"SUMMARY")
        print(f"{'='*80}")
        print(f"Total Deals Found: {len(deals)}")
        print(f"Output File: {args.output}")
        print(f"{'='*80}")
    else:
        print(f"\n{'='*80}")
        print("No deals found matching the criteria")
        print(f"{'='*80}")


if __name__ == '__main__':
    main()
